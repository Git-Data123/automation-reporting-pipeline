import pandas as pd
from pathlib import Path
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from src.transform import (
    transform_data,
    calculate_kpis,
    calculate_region_summary,
    calculate_product_summary,
)


def generate_report(kpis, region_summary, product_summary):
    """
    Generate a simple text report from KPI results.

    Args:
        kpis (dict): KPI summary values.
        region_summary (pd.DataFrame): Region-level summary table.
        product_summary (pd.DataFrame): Product-level summary table.

    Returns:
        str: Formatted report text.
    """
    report_lines = [
        "FINANCIAL REPORT",
        "================",
        "",
        f"Total Revenue : {kpis['total_revenue']}",
        f"Total Expenses: {kpis['total_expenses']}",
        f"Total Profit  : {kpis['total_profit']}",
        f"Total Records : {kpis['total_records']}",
        "",
        "Revenue / Expense / Profit by Region",
        "------------------------------------",
        region_summary.to_string(index=False),
        "",
        "Revenue / Expense / Profit by Product",
        "-------------------------------------",
        product_summary.to_string(index=False),
    ]

    report_text = "\n".join(report_lines)
    return report_text

def export_to_excel(transformed_df, kpis, region_summary, product_summary, output_file):
    """
    Export reporting outputs to a formatted Excel workbook.

    Args:
        transformed_df (pd.DataFrame): Transformed dataset.
        kpis (dict): KPI summary dictionary.
        region_summary (pd.DataFrame): Region summary table.
        product_summary (pd.DataFrame): Product summary table.
        output_file (str or Path): Output Excel file path.
    """
    kpi_df = pd.DataFrame([
        {"metric": key, "value": value}
        for key, value in kpis.items()
    ])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Write data to sheets
        transformed_df.to_excel(writer, sheet_name="Raw_Transformed_Data", index=False)
        kpi_df.to_excel(writer, sheet_name="KPI_Summary", index=False)
        region_summary.to_excel(writer, sheet_name="Region_Summary", index=False)
        product_summary.to_excel(writer, sheet_name="Product_Summary", index=False)

        # Create dashboard sheet
        workbook = writer.book
        dashboard_sheet = workbook.create_sheet("Dashboard")
        writer.sheets["Dashboard"] = dashboard_sheet

        # Dashboard title
        dashboard_sheet.merge_cells("A1:F1")
        dashboard_sheet["A1"] = "EXECUTIVE DASHBOARD"
        dashboard_sheet["A1"].font = Font(bold=True, size=18)
        dashboard_sheet["A1"].alignment = Alignment(horizontal="center")

        # KPI labels and values
        dashboard_sheet["A3"] = "Total Revenue"
        dashboard_sheet["B3"] = kpis["total_revenue"]

        dashboard_sheet["A4"] = "Total Expenses"
        dashboard_sheet["B4"] = kpis["total_expenses"]

        dashboard_sheet["A5"] = "Total Profit"
        dashboard_sheet["B5"] = kpis["total_profit"]

        dashboard_sheet["A6"] = "Total Records"
        dashboard_sheet["B6"] = kpis["total_records"]

        # Format KPI section
        for cell in ["A3", "A4", "A5", "A6"]:
            dashboard_sheet[cell].font = Font(bold=True)

        for cell in ["B3", "B4", "B5", "B6"]:
            dashboard_sheet[cell].font = Font(bold=True)
            dashboard_sheet[cell].number_format = "#,##0"

        for row in range(3, 7):
            dashboard_sheet[f"A{row}"].fill = PatternFill(
                start_color="E7F3FF",
                end_color="E7F3FF",
                fill_type="solid"
            )
            dashboard_sheet[f"B{row}"].fill = PatternFill(
                start_color="E7F3FF",
                end_color="E7F3FF",
                fill_type="solid"
            )

        dashboard_sheet.column_dimensions["A"].width = 20
        dashboard_sheet.column_dimensions["B"].width = 15
        dashboard_sheet.row_dimensions[2].height = 10

        # Header styles
        header_fill = PatternFill(
            start_color="C0C0C0",
            end_color="C0C0C0",
            fill_type="solid"
        )
        header_font = Font(bold=True)

        # Apply sheet-level formatting to table sheets only
        for sheet_name, worksheet in writer.sheets.items():
            if sheet_name == "Dashboard":
                continue

            worksheet.freeze_panes = "A2"

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    value = str(cell.value) if cell.value is not None else ""
                    if len(value) > max_length:
                        max_length = len(value)

                worksheet.column_dimensions[column_letter].width = max_length + 2

        # Format Raw_Transformed_Data sheet
        raw_sheet = writer.sheets["Raw_Transformed_Data"]
        for cell in raw_sheet["A"][1:]:
            cell.number_format = "YYYY-MM-DD"

        for column in ["D", "E", "F"]:
            for cell in raw_sheet[column][1:]:
                cell.number_format = "#,##0"

        # Format KPI_Summary sheet
        kpi_sheet = writer.sheets["KPI_Summary"]
        for cell in kpi_sheet["B"][1:]:
            cell.number_format = "#,##0"

        # Format Region_Summary sheet
        region_sheet = writer.sheets["Region_Summary"]
        for column in ["B", "C", "D"]:
            for cell in region_sheet[column][1:]:
                cell.number_format = "#,##0"

        # Format Product_Summary sheet
        product_sheet = writer.sheets["Product_Summary"]
        for column in ["B", "C", "D"]:
            for cell in product_sheet[column][1:]:
                cell.number_format = "#,##0"

        # Create chart for Region Summary
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Revenue by Region"
        chart1.y_axis.title = "Revenue"
        chart1.x_axis.title = ""
        chart1.legend = None
        chart1.width = 14
        chart1.height = 8

        region_data = Reference(region_sheet, min_col=2, min_row=1, max_row=region_sheet.max_row)
        region_categories = Reference(region_sheet, min_col=1, min_row=2, max_row=region_sheet.max_row)

        chart1.add_data(region_data, titles_from_data=True)
        chart1.set_categories(region_categories)
        chart1.y_axis.majorGridlines = None
        chart1.y_axis.tickLblPos = "nextTo"
        chart1.x_axis.tickLblPos = "nextTo"
        chart1.y_axis.scaling.min = 0

        dashboard_sheet.add_chart(chart1, "D2")

        # Create chart for Product Summary
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Revenue by Product"
        chart2.y_axis.title = "Revenue"
        chart2.x_axis.title = ""
        chart2.legend = None
        chart2.width = 14
        chart2.height = 8

        product_data = Reference(product_sheet, min_col=2, min_row=1, max_row=product_sheet.max_row)
        product_categories = Reference(product_sheet, min_col=1, min_row=2, max_row=product_sheet.max_row)

        chart2.add_data(product_data, titles_from_data=True)
        chart2.set_categories(product_categories)
        chart2.y_axis.majorGridlines = None
        chart2.y_axis.tickLblPos = "nextTo"
        chart2.x_axis.tickLblPos = "nextTo"
        chart2.y_axis.scaling.min = 0

        dashboard_sheet.add_chart(chart2, "D18")


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parent.parent
    file_path = project_root / "data" / "raw" / "sales_data.csv"

    df = pd.read_csv(file_path)

    transformed_df = transform_data(df)
    kpis = calculate_kpis(transformed_df)
    region_summary = calculate_region_summary(transformed_df)
    product_summary = calculate_product_summary(transformed_df)

    report = generate_report(kpis, region_summary, product_summary)
    print(report)